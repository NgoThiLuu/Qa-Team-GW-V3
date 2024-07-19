from cgitb import text
import time,datetime, json, random,os,sys,re,codecs
from os import pipe ,path
from xml.etree.ElementTree import Element
from openpyxl import load_workbook
from datetime import date,timedelta
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from selenium import webdriver
from colorama import Fore, Back, Style
from colorama import init, AnsiToWin32
from sys import platform
from pathlib import Path
from gw_v3_set_up import execution_log,fail_log,driver,xlsx_xpath,Param,File,data



class color():
    ENDC    = "\033[39m"
    PASS    = "\033[32m"
    FAIL    = "\033[31m"
    TITLE   = "\033[33m"
    CONTENT = "\033[37m"

class commons():
    def Time():
        Hour = Param.time
        Hour = Hour[None: int(Hour.rfind("."))]
        return data["date_hour"] %(commons.Today(), Hour)

    def Url(domain):
        return data["url"] % domain
    
    def Menu(domain):
        return data["menu"] % domain

    def Today():
        return str(datetime.date.today()).replace(" ","")

    def Logging(text):
        log_msg =  codecs.open(execution_log, "a" ,"utf-8")
        log_msg.write(str(text) + "\n")
        log_msg.close()

    def ValidateFailResultAndSystem(fail_msg):
        append_fail_result =  codecs.open(fail_log, "a" ,"utf-8")
        append_fail_result.write(data["FailCase"] % str(fail_msg))
        #append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
        append_fail_result.close()
    
    def Title(content) :
        commons.Logging(content)
        print(color.TITLE + content + color.ENDC)

    def CasePass(content) :
        commons.Logging(content)
        print(color.PASS  + data["PASS"] % content + color.ENDC)

    def CaseFail(content) :
        commons.ValidateFailResultAndSystem(content)
        print(color.FAIL  + data["FAIL"] % content  + color.ENDC)

    def Content(content):
        commons.Logging(content)
        print(color.CONTENT + content + color.ENDC)

    def FindElementById(id):
        driver.implicitly_wait(10)
        element = driver.find_element(By.ID,id)
        return element

    def ReturnElement(xpath) :
        element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,xpath)))
        return element
        
    def ClickElementWithXpath(xpath):
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,xpath))).click()

    def ClickElementWithText(text):
        WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.LINK_TEXT,text))).click()

    def IsDisplayedByXpath(xpath):
        try:
            time.sleep(5)
            driver.find_element("xpath",xpath)
            #driver.find_element_by_xpath(xpath)
            return True
            
        except NoSuchElementException:
            return False
    
    def IsDisplayedByCss(xpath):
        try:
            time.sleep(5)
            #driver.find_element("xpath",xpath)
            driver.find_element_by_css_selector(xpath)
            return True
            
        except NoSuchElementException:
            return False
        
    def IsDisplayedByTextLink(text):
        try:
            driver.find_element_by_link_text(text)
            return True
        except NoSuchElementException:
            return False
        
    def IsDisplayedById(id):
        try:
            WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID,id)))
            return True
        except NoSuchElementException:
            return False
    
    def TotalData(list):
        total = 0
        for element in list :
            total = total + 1
        return total

    def ClickLinkText(text):
        driver.find_element_by_link_text(text).click()
        driver.implicitly_wait(5)

    def SwitchToFrame(id_frame , id_element , key):
        driver.switch_to.frame(commons.FindElementById(id_frame))
        commons.FindElementById(id_element).send_keys(key)
       
    def SwitchToDefaultContent():
        driver.switch_to.default_content()
    
    def Scroll():
        html = driver.find_element_by_tag_name('html')
        html.send_keys(Keys.END)

    def ScrollingToTarget(target):
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_to_element(target).perform()

    def ChangeLanguage():
        try :
            time.sleep(3)
            #driver.find_element_by_xpath(data["ava"]).click()
            User       = driver.find_element_by_xpath(data["ava"])
            Text       = User.text
            Department = Text[None: int(Text.rfind("\n"))]
            Text       = Text[int(Text.rfind("\n")) + 1 : None]
            Name       = Text[None: int(Text.rfind(" "))]
            Position   = Text[int(Text.rfind(" ")) : None]
            Info_User  = {"Department":Department,"Name":Name ,"Position":Position}
            User.click()
            driver.find_element_by_xpath(data["ic_setting"]).click()
            driver.find_element_by_xpath(data["sl_lang"]).click()
            time.sleep(1)
            if commons.IsDisplayedByXpath(data["en"]) == True :
                commons.CasePass("Current language is Eng")
            else:
                driver.find_element_by_xpath(data["sl_en"]).click()
                time.sleep(1)
                driver.find_element_by_xpath(data["ok"]).click()
                commons.CasePass("Current language is Eng")

            return Info_User
        except:
            return False
    
    def SendKey(key,xpath):
        time.sleep(5)
        Input_User = driver.find_element_by_xpath(xpath)
        Input_User.clear()
        Input_User.click()
        Input_User.send_keys(key)
        Input_User.send_keys(Keys.RETURN)
        time.sleep(5)
        if  str(Input_User.get_attribute('value')) == str(key) :
            return True
        else :
            return False

    def GetTextWithI(xpath , i):
        text = driver.find_element_by_xpath(xpath % str(i)).text 
        return text

    def GetText(xpath):
        text = driver.find_element_by_xpath(xpath).text 
        return text

    def ReplaceSpace(text):
        return text.replace(" ", "")

    def ReplaceCustom(text,content):
        return text.replace(content, "")

    def AddData(List,Value):
        List["after"] = Value
        return List
    
    def AddTotal(List,Value):
        List["Total_Request"] = Value
        return List
    
    def AddPar(List,Value,Key):
        List[Key] = Value
        return List

    def WaitToClick(xpath):
        time.sleep(3)
        driver.find_element_by_xpath(xpath).click()

    def WriteOnExcel(content_excel):
        if content_excel["status"] == "Pass":
            commons.Logging(content_excel["description"])
            print(color.PASS + data["PASS"] % content_excel["description"] + color.ENDC)
        else :
            commons.ValidateFailResultAndSystem(content_excel["description"])
            print(color.FAIL + data["FAIL"] % content_excel["description"] + color.ENDC)

        wb = openpyxl.load_workbook(xlsx_xpath) 
        if content_excel["sheet"] == "ac":
            sheet_use = wb.get_sheet_by_name(Param.access_page)
        else:
            sheet_use = wb.get_sheet_by_name(Param.functions) 
        row = sheet_use.max_row
        col = sheet_use.max_column
       
        if content_excel["status"] == "Fail":
            sheet_use.cell(row+1,col-6).font= Font(color='FF0000')
            sheet_use.cell(row+1,col-5).font= Font(color='FF0000')
            sheet_use.cell(row+1,col-4).font= Font(color='FF0000')
            sheet_use.cell(row+1,col-3).font= Font(color='FF0000')
            sheet_use.cell(row+1,col-2).font= Font(color='FF0000')
            sheet_use.cell(row+1,col-1).font= Font(color='FF0000')
            sheet_use.cell(row+1,col).font= Font(color='FF0000')

        sheet_use.cell(row=row+1,column=col-6).value=content_excel["menu"]
        sheet_use.cell(row=row+1,column=col-5).value=content_excel["submenu"]
        sheet_use.cell(row=row+1,column=col-4).value=content_excel["testcase"]
        sheet_use.cell(row=row+1,column=col-3).value=content_excel["status"]
        sheet_use.cell(row=row+1,column=col-2).value=content_excel["description"]
        sheet_use.cell(row=row+1,column=col-1).value=File.date_xls
        sheet_use.cell(row=row+1,column=col).value=content_excel["tester"]
        wb.save(xlsx_xpath)
    





