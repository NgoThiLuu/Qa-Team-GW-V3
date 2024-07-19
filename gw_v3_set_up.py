import time,datetime, json, random,os,sys,re,codecs
from os import pipe ,path
from openpyxl import load_workbook
from datetime import date
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from selenium import webdriver
from colorama import Fore, Back, Style
from colorama import init, AnsiToWin32
from sys import platform
from pathlib import Path
'''
f = open('nul', 'w')
sys.stderr= f
'''
def Replace(text):
    return text("\\","/")

class Param():
    access_page = "Access Page"
    functions   = "Functions"
    json_name   = "\\gw_v3_data.json"
    xlsx_name   = "MenuMailArchiveContactBoard_"
    format_xlsx = ".xlsx"
    format_txt  = ".txt"
    local_log   = "\\Log\\execution_log_"
    log_folder  = "\\Log\\"
    chrome      = "\\chromedriver.exe"
    time        = str(datetime.datetime.now().time())
    local       = os.path.dirname(Path(__file__).absolute())
    now         = datetime.datetime.now()
    date_time   = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id     = date_time.replace("/","").replace(", ","").replace(":", "")[2:]

class File():
    # Color #
    init(wrap = False)
    stream    = AnsiToWin32(sys.stderr).stream

    # Json #
    json_file = Param.local + Param.json_name

    # Excel #
    XlsxName  = Param.xlsx_name + str(Param.date_id) + Param.format_xlsx
    date_xls  = str(datetime.date.today()) + "," + Param.time[None:Param.time.rfind(".")]

    # log #
    execution_log = Param.local + Param.local_log + str(Param.date_id) + Param.format_txt
    fail_log      = execution_log.replace("execution_log_", "fail_log_")
    error_log     = execution_log.replace("execution_log_", "error_log_")
        
def CreateXlsx(data):
    title      = data["title"]
    sheet_name = [Param.functions,Param.access_page]
    wb         = openpyxl.Workbook()
    wb.save(xlsx_xpath)  
    wb1        = load_workbook(xlsx_xpath)
    for name in sheet_name :
        wb1.create_sheet(name)
        ws1 = wb1.get_sheet_by_name(name)
        for i in range(1,8):
            ws1.cell(row=1,column=i).value=title[str(i)]
        col = ws1.max_column
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['E'].width = 60
        ws1.column_dimensions['F'].width = 20
        ws1.column_dimensions['G'].width = 15
        my_red  = openpyxl.styles.colors.Color(rgb='00103667')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = my_red)
        for col in range(1,col+1):
            ws1.cell(1,col,value=None).alignment = Alignment(horizontal='center')
            ws1.cell(1,col,value=None).font= Font(size=12, color='FFFFFF', bold=True)
            ws1.cell(1,col,value=None).fill=my_fill
        wb1.save(xlsx_xpath)  
    sh = wb1.get_sheet_by_name('Sheet')
    wb1.remove_sheet(sh)
    wb1.save(xlsx_xpath)



if platform == "linux" or platform == "linux2":
    execution_log = Replace(File.execution_log)
    json_file     = Replace(File.json_file)
    fail_log      = Replace(File.fail_log)
    error_log     = Replace(File.error_log)
    xlsx_xpath    = Param.local + Replace(Param.log_folder) + File.XlsxName
    driver        = webdriver.Chrome("/usr/bin/chromedriver")
    driver.maximize_window()

else :
    json_file     =  File.json_file
    execution_log = File.execution_log
    fail_log      = File.fail_log
    error_log     = File.error_log
    xlsx_xpath    =  Param.local + Param.log_folder + File.XlsxName
    driver_path   = Param.local + Param.chrome
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(driver_path , chrome_options = chrome_options)
    driver.maximize_window()
  
with open(json_file) as json_file:
    data = json.load(json_file)

# create log file of fail test case
open(execution_log, "x").close()

# create log file of fail test case
open(fail_log, "x").close()

# create log file of fail test case
open(error_log, "x").close()

# create xlsx file to write test case
CreateXlsx(data)






